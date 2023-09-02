import random
from functools import lru_cache

import openpyxl
from django.contrib.auth import logout, login
from django.contrib.auth.views import LoginView
from django.core.mail import send_mail
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse, reverse_lazy
from django.utils.crypto import get_random_string
from django.views import View
from django.views.generic import TemplateView
from rest_framework.response import Response
from rest_framework.views import APIView

from tsartvie import settings
from .forms import LoginUserForm
from .models import BinaryDict, HandWriting, Question, Answer, UserAnswer, Text, User, Message, Balance, TypeQuestion, \
    Code, Verdict, Commandment, Role


class MainPageView(View):

    def get(self, request, *args, **kwargs):
        return render(request, "index.html", context={
            'index': True,
            'verdicts': Verdict.objects.all(),
            'commandments': Commandment.objects.all()
        })


class VerdictShowView(View):

    def get(self, request, *args, **kwargs):
        return render(request, 'verdict.html', context={
            'verdict': Verdict.objects.get(id=kwargs['id']),
            'index': True,
        })


def logout_user(request):
    """Log out"""
    logout(request)
    return redirect('index')


class UserRegisterView(View):
    LINE_SYMBOL = '1234567890-qwertyuiop[]asdfghjkl;zxcvbnm,./~!@#$%^&*()_+QWERTYUIOP{}ASDFGHJKL:"ZXCVBNM<>?'
    LIST_REQUIRED_SYMBOLS = ["", "'", "-", " ", "!", "#", "$", "%", "&", "(", ")", "*", ",", ".", "/", ":", ";",
                             "?", "@", "[", "]", "^", "_", "{", "|", "}", "~", "+", "<", ">", "0", "1", "2", "3",
                             "4", "5", "6", "7", "8", "9", "a", "A", "b", "B", "c", "C", "d", "D", "e", "E", "f", "F",
                             "g", "G", "h", "H", "i", "I", "j", "J", "k", "K", "l", "L", "m", "M", "n", "N", "o", "O",
                             "p", "P", "q", "Q", "r", "R", "s", "S", "t", "T", "u", "U", "v", "V", "w", "W", "x", "X",
                             "y", "Y", "z", "Z", "а", "А", "б", "Б", "в", "В", "г", "Г", "д", "Д", "е", "Е", "ё", "Ё",
                             "ж", "Ж", "з", "З", "и", "И", "й", "Й", "к", "К", "л", "Л", "м", "М", "н", "Н", "о", "О",
                             "п", "П", "р", "Р", "с", "С", "т", "Т", "у", "У", "ф", "Ф", "х", "Х", "ц", "Ц", "ч", "Ч",
                             "ш", "Ш", "щ", "Щ", "ъ", "Ъ", "ы", "Ы", "ь", "Ь", "э", "Э", "ю", "Ю", "я", "Я"]

    @classmethod
    def random_hex(cls):
        result = ""
        for i in range(30):
            result += cls.LINE_SYMBOL[random.randint(0, len(cls.LINE_SYMBOL) - 1)]
        return result

    @staticmethod
    def get_client_ip(request):  # ф-ция определяющая ip адресс с которого пробросили запрос
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')  # request отвечает за данные с запроса
        if x_forwarded_for:  # мета отвечает за данные в запросе
            ip = x_forwarded_for.split(',')[0].strip()  # разбивает ip адресс по запятой и берёт 1й элемент
        else:  # а также удаляет пробелы до и после строки
            ip = request.META.get('REMOTE_ADDR')
        return ip  # возращает ip адресс из ф-ции

    @staticmethod
    @lru_cache
    def _code(email, code):
        try:
            return Code.objects.get(email=email, code=code)
        except Code.DoesNotExist:
            return None

    def get(self, request, *args, **kwargs):
        return render(request, 'register.html')

    def post(self, request, *args, **kwargs):
        user_email = request.session.get('user_email')
        code_user = request.POST.get("code")
        if not code_user or not self._code(user_email, code_user):
            return redirect('register')
        user = User.objects.create(
            email=user_email,
            password=request.session.get('user_pass'),
            username=get_random_string(12, self.LINE_SYMBOL)
        )
        user.set_password(request.session.get('user_pass'))
        user.ip_address = self.get_client_ip(self.request)
        user.save()
        self._code(user_email, code_user).delete()
        del request.session['user_email']
        del request.session['user_pass']
        login(self.request, user)
        for symbol in self.LIST_REQUIRED_SYMBOLS:
            binary = ""
            performance = self.random_hex()
            for symbol_performance in performance:
                binary += BinaryDict.objects.get(symbol=symbol_performance).binary
            HandWriting.objects.create(
                symbol=symbol,
                performance=performance,
                binary=binary,
                user=user
            )
        Balance.objects.create(user=user)
        return redirect('profile')


class UserLoginView(LoginView):
    form_class = LoginUserForm
    template_name = 'login.html'

    def get_success_url(self):
        return reverse_lazy('studying')


class VoteShowView(View):
    def get(self, request, *args, **kwargs):
        return render(request, "vote.html", context={
            'questions': Question.objects.filter(
                questions_user__answer__isnull=True,
                type_question__id=kwargs['id']
            ),
            'type_question': kwargs['id']
        })

    def post(self, request, *args, **kwargs):
        question = Question.objects.get(id=request.POST['question_id'])
        answer = Answer.objects.get(id=request.POST['answer_id'])
        UserAnswer.objects.create(
            question=question,
            answer=answer,
            user=request.user,
        )
        return redirect(reverse('questions', kwargs={'id': kwargs['id']}))


class VoteListCreateView(View):

    def get(self, request, *args, **kwargs):
        return render(request, "type_question.html", context={
            'type_questions': TypeQuestion.objects.all()
        })


class UploadSymbolView(View):

    def post(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="hand_writing.xlsx"'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Hand Writing Data'
        headers = ['Символ', 'Шестнадцатеричное', 'Бинарь']
        worksheet.append(headers)
        hand_writing = HandWriting.objects.filter(user=request.user)
        for symbol in hand_writing:
            data_row = [symbol.symbol, symbol.performance, symbol.binary]
            worksheet.append(data_row)

        workbook.save(response)
        return response


class TextCreateView(View):

    def post(self, request, *args, **kwargs):
        Text.objects.create(
            text=request.POST['text'],
            user=request.user
        )
        return redirect('work')


class UserProfileView(View):

    def get(self, request, *args, **kwargs):
        return render(request, 'profile.html', context={
            'texts': Text.objects.filter(user=request.user).order_by('-created_at')
        })


class TextUploadView(View):

    def post(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="hand_writing.xlsx"'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Выгрузка текста'
        headers = ['Символ', 'Шестнадцатеричное', 'Бинарь']
        worksheet.append(headers)
        text = Text.objects.get(id=kwargs['id']).text
        performance = "".join([self._hand_writings[symbol][0] for symbol in text])
        binary = "".join([self._hand_writings[symbol][1] for symbol in text])
        worksheet.append([text, performance, binary])
        workbook.save(response)
        return response

    @property
    @lru_cache
    def _hand_writings(self):
        hand_writings = HandWriting.objects.filter(
            user=self.request.user
        )
        return {
            hand_writing.symbol: [hand_writing.performance, hand_writing.binary]
            for hand_writing in hand_writings
        }


class TextDeleteView(View):

    def post(self, request, *args, **kwargs):
        Text.objects.get(
            id=kwargs['id'],
            user=request.user
        ).delete()
        return redirect('profile')


class VoteListResultView(View):

    def get(self, request, *args, **kwargs):
        user_answers = UserAnswer.objects.annotate(

        ).filter(user=request.user)
        return render(request, 'vote_result.html', context={
            'user_answers': user_answers
        })


class EmailShowView(View):

    def get(self, request, *args, **kwargs):
        return render(request, 'mail.html', context={
            'users': User.objects.filter(~Q(id=request.user.id)),
            'messages': Message.objects.filter(user=request.user)
        })

    def post(self, request, *args, **kwargs):
        text = request.POST.get("text_message")
        if not text:
            return redirect('email')
        Message.objects.create(
            text=text,
            user=User.objects.get(id=request.POST['user_id']),
            author=request.user
        )
        return redirect('email')


class MessageUploadView(View):

    def post(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="message.xlsx"'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Выгрузка сообщения'
        headers = ['Символ', 'Шестнадцатеричное', 'Бинарь']
        worksheet.append(headers)
        message = Message.objects.get(id=kwargs['id']).text
        performance = "".join([self._hand_writings[symbol][0] for symbol in message])
        binary = "".join([self._hand_writings[symbol][1] for symbol in message])
        worksheet.append([message, performance, binary])
        workbook.save(response)
        return response

    @property
    @lru_cache
    def _hand_writings(self):
        hand_writings = HandWriting.objects.filter(
            user=self.request.user
        )
        return {
            hand_writing.symbol: [hand_writing.performance, hand_writing.binary]
            for hand_writing in hand_writings
        }


class UserFinanceView(View):

    def get(self, request, *args, **kwargs):
        return render(request, 'finance.html', context={
            'balance': Balance.objects.get(user=request.user).balance
        })


class UserKeyView(View):

    def get(self, request, *args, **kwargs):
        return render(request, 'keys.html', context={

        })


class EmailSentShowView(View):
    def get(self, request, *args, **kwargs):
        return render(request, 'list_sent_user.html', context={
            'sent_messages': Message.objects.filter(author=request.user)
        })


class StudyingShowView(View):

    def get(self, request, *args, **kwargs):
        return render(request, 'studying.html', context={
        })


class WorkShowView(View):

    def get(self, request, *args, **kwargs):
        return render(request, 'work.html', context={
            'texts': Text.objects.filter(user=request.user).order_by('-created_at')
        })


class MatrixPage(TemplateView):
    template_name = 'matrix.html'


class UserVerifyView(View):

    def get(self, request, *args, **kwargs):
        return render(request, "verify.html", context={
            'status': Role.objects.all()
        })

    def post(self, request, *args, **kwargs):
        email = request.POST.get('email')
        password = request.POST.get('pass')
        if not email or not password:
            return redirect('verify')
        code = get_random_string(6, '0123456789')
        Code.objects.create(
            email=email,
            code=code
        )
        self.request.session['user_email'] = email
        self.request.session['user_pass'] = password
        send_mail(
            subject='Верификация почты',
            message=f'Для верификации почты введите данный код {code}',
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=[email],
            fail_silently=True,
        )
        return redirect('register')


class TextCompilView(APIView):
    @property
    @lru_cache
    def _simple(self):
        return {
            hand_writing.symbol: [hand_writing.performance, hand_writing.binary]
            for hand_writing in HandWriting.objects.filter(user=self.request.user)
        }

    @property
    @lru_cache
    def _hexadecimal(self):
        return {
            hand_writing.performance: [hand_writing.symbol, hand_writing.binary]
            for hand_writing in HandWriting.objects.filter(user=self.request.user)
        }

    @property
    @lru_cache
    def _binary(self):
        return {
            hand_writing.binary: [hand_writing.symbol, hand_writing.performance]
            for hand_writing in HandWriting.objects.filter(user=self.request.user)
        }

    @property
    @lru_cache
    def _binary_model(self):
        return {
            item.binary: item.symbol
            for item in BinaryDict.objects.all()
        }

    def get(self, request):
        type_text = request.query_params.get('text_type')
        text = request.query_params.get('text')
        if not type_text or not text:
            return Response()

        if type_text == "simple":
            return Response(
                {
                    "result": dict(
                        simple=text,
                        hexadecimal="".join([self._simple[symbol][0] for symbol in text]),
                        binary="".join([self._simple[symbol][1] for symbol in text])
                    )
                }
            )
        elif type_text == "hexadecimal":
            list_hexadecimal = [text[i:i + 30] for i in range(0, len(text), 30)]
            return Response(
                {
                    "result": dict(
                        simple="".join([self._hexadecimal[item][0] for item in list_hexadecimal]),
                        hexadecimal=text,
                        binary="".join([self._hexadecimal[item][1] for item in list_hexadecimal]),
                    )
                }
            )
        elif type_text == "binary":
            list_binary = [text[i:i + 12] for i in range(0, len(text), 12)]  # [0101010010101, ...]
            hexadecimals = "".join([self._binary_model[item] for item in list_binary])  # [12,pdfsfdp,324, ...]
            list_hexadecimal = [hexadecimals[i:i + 30] for i in range(0, len(hexadecimals), 30)]
            return Response(
                {
                    "result": dict(
                        simple="".join([self._hexadecimal[item][0] for item in list_hexadecimal]),
                        hexadecimal="".join(list_hexadecimal),
                        binary=text,
                    )
                }
            )
        return Response({})
